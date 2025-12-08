from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, abort
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime
from . import db
from .models import Requirement, RequirementVersion, Project, ProjectFile
from .services.ai_client import AIClient, generate_new_requirements, optimize_excel_requirements
from .services.exel_service import parse_excel_to_data

agent_bp = Blueprint('agent', __name__, url_prefix='/agent')


def normalize_key(title: str) -> str:
    """
    Normalize a requirement title to create a unique key.
    
    Args:
        title (str): The requirement title
    
    Returns:
        str: Normalized key (lowercase, no special chars, max 100 chars)
    """
    import re
    # Convert to lowercase and replace spaces with underscores
    key = title.lower().strip()
    # Remove special characters except underscores and hyphens
    key = re.sub(r'[^a-z0-9_\-]', '_', key)
    # Replace multiple underscores with single underscore
    key = re.sub(r'_+', '_', key)
    # Trim to max 100 characters
    key = key[:100]
    # Remove leading/trailing underscores
    key = key.strip('_')
    return key


@agent_bp.route('/<int:project_id>')
@login_required
def agent_page(project_id):
    """Render the AI agent page for a specific project"""
    project = Project.query.get_or_404(project_id)
    
    # Authorization check
    if project.user_id != current_user.id:
        abort(403)
    
    return render_template('agent/agent.html', project=project)


@agent_bp.route('/upload/<int:project_id>')
@login_required
def upload_page(project_id):
    """Render a dedicated upload page where users can upload an Excel file to be
    analyzed/optimized by the AI."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

    return render_template('agent/upload.html', project=project)


@agent_bp.route('/upload_excel/<int:project_id>', methods=['POST'])
@login_required
def upload_excel_route(project_id):
    """Upload Excel file and import requirements directly WITHOUT AI processing"""
    project = Project.query.get_or_404(project_id)
    
    # Authorization check
    if project.user_id != current_user.id:
        abort(403)
    
    try:
        # Check if Excel file was uploaded
        if 'excel_file' not in request.files:
            return jsonify({
                'ok': False,
                'error': 'Keine Datei hochgeladen.'
            }), 400
        
        file = request.files['excel_file']
        if not file or not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'ok': False,
                'error': 'Bitte eine gültige Excel-Datei (.xlsx oder .xls) hochladen.'
            }), 400
        
        # Secure filename and add timestamp
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{name}_{timestamp}{ext}"

        # Save file permanently
        uploads_dir = os.path.join('uploads')
        os.makedirs(uploads_dir, exist_ok=True)
        uploaded_file_path = os.path.join(uploads_dir, unique_filename)
        file.save(uploaded_file_path)

        try:
            # Parse Excel file
            excel_data = parse_excel_to_data(uploaded_file_path)
            
            # Remove system columns (Version, ID)
            cleaned_excel_data = []
            for row in excel_data:
                cleaned_row = {k: v for k, v in row.items() if k.lower() not in ['version', 'id']}
                if cleaned_row:
                    cleaned_excel_data.append(cleaned_row)
            
            if not cleaned_excel_data:
                return jsonify({
                    'ok': False,
                    'error': 'Keine Daten in der Excel-Datei gefunden.'
                }), 400

            # Extract and merge column names
            excel_columns = list(cleaned_excel_data[0].keys())
            filtered_columns = [col for col in excel_columns if col and col.strip()]
            
            if filtered_columns:
                existing_columns = project.get_custom_columns() or []
                merged_columns = existing_columns.copy()
                for col in filtered_columns:
                    if col not in merged_columns:
                        merged_columns.append(col)
                
                project.set_custom_columns(merged_columns)
                db.session.commit()

            # Create ProjectFile entry
            project_file = ProjectFile(
                project_id=project_id,
                filename=filename,
                filepath=uploaded_file_path,
                file_type='upload',
                created_by_id=current_user.id
            )
            db.session.add(project_file)
            db.session.commit()
            uploaded_file_id = project_file.id

            # Get optional user description from form
            user_description = request.form.get('user_description', '').strip()

            # Optimize requirements using AI while maintaining structure
            try:
                optimized_reqs = optimize_excel_requirements(
                    existing_requirements=cleaned_excel_data,
                    columns=filtered_columns if filtered_columns else ["title", "description", "category"],
                    user_description=user_description if user_description else None
                )
            except Exception as e:
                return jsonify({
                    'ok': False,
                    'error': f'KI-Optimierung fehlgeschlagen: {str(e)}'
                }), 500

            if not optimized_reqs:
                return jsonify({
                    'ok': False,
                    'error': 'Keine optimierten Anforderungen generiert.'
                }), 400

            # Import optimized requirements
            saved_count = 0
            for req_data in optimized_reqs:
                # Find title
                title = None
                for title_key in ['title', 'Title', 'titel', 'Titel', 'name', 'Name']:
                    if title_key in req_data:
                        title = req_data.get(title_key, '').strip()
                        if title:
                            break
                
                if not title:
                    for key, value in req_data.items():
                        if value and str(value).strip():
                            title = str(value).strip()
                            break
                
                # Find description
                description = None
                for desc_key in ['description', 'Description', 'beschreibung', 'Beschreibung', 'text', 'Text']:
                    if desc_key in req_data:
                        description = req_data.get(desc_key, '').strip()
                        if description:
                            break
                
                if not description:
                    values = [str(v).strip() for v in req_data.values() if v and str(v).strip() and str(v).strip() != title]
                    description = values[0] if values else "Keine Beschreibung"
                
                if not title:
                    continue
                
                # Create normalized key
                key = normalize_key(title)
                
                # Check if requirement exists
                existing_req = Requirement.query.filter_by(project_id=project_id, key=key).first()
                
                if not existing_req:
                    req = Requirement(project_id=project_id, key=key)
                    db.session.add(req)
                    db.session.flush()
                    version_index = 1
                    version_label = 'A'
                else:
                    req = existing_req
                    db.session.refresh(req)
                    
                    if req.versions:
                        max_version_index = max(v.version_index for v in req.versions)
                        version_index = max_version_index + 1
                    else:
                        version_index = 1
                    version_label = chr(ord('A') + (version_index - 1))
                
                # Get category
                category = ''
                for cat_key in ['category', 'Category', 'kategorie', 'Kategorie', 'cat', 'Cat']:
                    if cat_key in req_data:
                        category = req_data.get(cat_key, '').strip()
                        if category:
                            break
                
                # Get status
                status = 'Offen'
                for status_key in ['status', 'Status']:
                    if status_key in req_data:
                        status_val = req_data.get(status_key, '').strip()
                        if status_val in ['Offen', 'In Arbeit', 'Fertig']:
                            status = status_val
                            break
                
                # Create version
                new_version = RequirementVersion(
                    requirement_id=req.id,
                    version_index=version_index,
                    version_label=version_label,
                    title=title,
                    description=description,
                    category=category,
                    status=status,
                    created_by_id=current_user.id,
                    source_file_id=uploaded_file_id
                )
                
                # Store custom data
                custom_data = {}
                for col, value in req_data.items():
                    if col.lower() not in ['version', 'id']:
                        if value and str(value).strip():
                            custom_data[col] = str(value).strip()
                
                if custom_data:
                    new_version.set_custom_data(custom_data)
                
                db.session.add(new_version)
                db.session.flush()
                saved_count += 1
            
            db.session.commit()

            redirect_url = url_for('main.project_overview', project_id=project_id, file_id=uploaded_file_id)
            
            return jsonify({
                'ok': True,
                'count': saved_count,
                'redirect': redirect_url,
                'message': f'{saved_count} Anforderungen aus Excel importiert und mit KI optimiert.'
            })

        except Exception as e:
            # Clean up on error
            if uploaded_file_path and os.path.exists(uploaded_file_path):
                try:
                    os.remove(uploaded_file_path)
                except:
                    pass
            raise e
            
    except Exception as e:
        return jsonify({
            'ok': False,
            'error': f'Fehler beim Hochladen: {str(e)}'
        }), 500


@agent_bp.route('/generate/<int:project_id>', methods=['POST'])
@login_required
def generate_requirements_route(project_id):
    """Generate requirements using AI for a specific project"""
    project = Project.query.get_or_404(project_id)
    
    # Authorization check
    if project.user_id != current_user.id:
        abort(403)
    
    try:
        # Get form data
        user_description = request.form.get('user_description', '').strip()
        inputs_json = request.form.get('inputs', '[]')
        new_columns_json = request.form.get('new_columns', '[]')
        
        # Parse inputs
        try:
            inputs_list = json.loads(inputs_json)
            inputs = {item['key']: item['value'] for item in inputs_list if 'key' in item and 'value' in item}
        except:
            inputs = {}
        
        # Parse new columns
        try:
            new_columns = json.loads(new_columns_json)
            if new_columns and isinstance(new_columns, list):
                # Merge new columns with existing project columns
                existing_columns = project.get_custom_columns() or []
                merged_columns = existing_columns.copy()
                for col in new_columns:
                    if col and col.strip() and col not in merged_columns:
                        merged_columns.append(col)
                
                # Update project columns
                if len(merged_columns) > len(existing_columns):
                    project.set_custom_columns(merged_columns)
                    db.session.commit()
        except Exception as e:
            pass  # Silently ignore column parsing errors
        
        # This endpoint is ONLY for AI generation (no Excel upload)
        # For Excel upload, use /upload_excel endpoint instead
        
        # Get project's custom columns
        custom_columns = project.get_custom_columns()
        columns = custom_columns if custom_columns else ["title", "description", "category"]
        
        # Generate requirements using AI (always use AI, no direct import)
        generated_reqs = generate_new_requirements(
            user_description=user_description if user_description else None,
            inputs=inputs,
            columns=columns
        )
        
        if not generated_reqs:
            return jsonify({
                'ok': False,
                'error': 'Keine Requirements generiert. Bitte versuchen Sie es erneut.'
            }), 400
        
        # Create the Excel snapshot file FIRST
        generated_file_id = None
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Generated Requirements"

            # use same columns list
            headers = [c.capitalize() for c in columns]
            # write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # write generated rows
            row_num = 2
            for req in generated_reqs:
                row = []
                for col in columns:
                    row.append(req.get(col, ""))
                for col_num, val in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row_num += 1

            # set some column widths
            for i in range(1, len(headers) + 1):
                col_letter = chr(ord('A') + i - 1)
                ws.column_dimensions[col_letter].width = 20

            # save workbook to uploads dir
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"generated_requirements_{project.name.replace(' ', '_')}_{timestamp}.xlsx"
            uploads_dir = os.path.join('uploads')
            os.makedirs(uploads_dir, exist_ok=True)
            filepath = os.path.join(uploads_dir, filename)
            wb.save(filepath)

            # create ProjectFile entry
            project_file = ProjectFile(
                project_id=project_id,
                filename=filename,
                filepath=filepath,
                file_type='generated',
                created_by_id=current_user.id
            )
            db.session.add(project_file)
            db.session.flush()  # Flush to get the ID
            
            # Get the file ID for linking requirements
            generated_file_id = project_file.id
        except Exception as e:
            # Non-fatal: silently ignore snapshot errors
            generated_file_id = None
        
        # Now save generated requirements to database with correct source_file_id
        saved_count = 0
        for req_data in generated_reqs:
            # Try to find title and description from various column names
            # Try common variations for title
            title = None
            for title_key in ['title', 'Title', 'titel', 'Titel', 'name', 'Name']:
                if title_key in req_data:
                    title = req_data.get(title_key, '').strip()
                    if title:
                        break
            
            # If no title found, use first non-empty value
            if not title:
                for key, value in req_data.items():
                    if value and str(value).strip():
                        title = str(value).strip()
                        break
            
            # Try common variations for description
            description = None
            for desc_key in ['description', 'Description', 'beschreibung', 'Beschreibung', 'text', 'Text']:
                if desc_key in req_data:
                    description = req_data.get(desc_key, '').strip()
                    if description:
                        break
            
            # If no description found, use second non-empty value or concatenate all
            if not description:
                values = [str(v).strip() for v in req_data.values() if v and str(v).strip() and str(v).strip() != title]
                description = values[0] if values else "Keine Beschreibung"
            
            if not title:
                continue
            
            # Create normalized key
            key = normalize_key(title)
            
            # Check if requirement with this key already exists
            existing_req = Requirement.query.filter_by(project_id=project_id, key=key).first()
            
            if not existing_req:
                # Create new requirement
                req = Requirement(project_id=project_id, key=key)
                db.session.add(req)
                db.session.flush()
                version_index = 1
                version_label = 'A'
            else:
                # Requirement exists - add new version
                req = existing_req
                
                # Refresh the requirement to get latest versions (in case we added one in this loop)
                db.session.refresh(req)
                
                # Find the highest version_index across ALL versions
                if req.versions:
                    max_version_index = max(v.version_index for v in req.versions)
                    version_index = max_version_index + 1
                else:
                    version_index = 1
                version_label = chr(ord('A') + (version_index - 1))
            
            # Create requirement version
            # Try to get category from various column names
            category = ''
            for cat_key in ['category', 'Category', 'kategorie', 'Kategorie', 'cat', 'Cat']:
                if cat_key in req_data:
                    category = req_data.get(cat_key, '').strip()
                    if category:
                        break
            
            # Try to get status from various column names
            status = 'Offen'
            for status_key in ['status', 'Status']:
                if status_key in req_data:
                    status_val = req_data.get(status_key, '').strip()
                    if status_val in ['Offen', 'In Arbeit', 'Fertig']:
                        status = status_val
                        break
            
            new_version = RequirementVersion(
                requirement_id=req.id,
                version_index=version_index,
                version_label=version_label,
                title=title,
                description=description,
                category=category,
                status=status,
                created_by_id=current_user.id,
                source_file_id=generated_file_id  # Only generated files in this endpoint
            )
            
            # Store ALL data from req_data as custom data (excluding system columns)
            custom_data = {}
            for col, value in req_data.items():
                # Skip system-managed columns
                if col.lower() in ['version', 'id']:
                    continue
                # Only store non-empty values
                if value and str(value).strip():
                    custom_data[col] = str(value).strip()
            
            if custom_data:
                new_version.set_custom_data(custom_data)
            
            db.session.add(new_version)
            db.session.flush()  # Flush immediately so next iteration sees this version
            saved_count += 1
        
        db.session.commit()

        # Redirect to the generated requirements
        if generated_file_id:
            redirect_url = url_for('main.project_overview', project_id=project_id, file_id=generated_file_id)
        else:
            redirect_url = url_for('main.project_overview', project_id=project_id)

        return jsonify({
            'ok': True,
            'count': saved_count,
            'redirect': redirect_url,
            'message': f'{saved_count} Anforderungen mit KI generiert.'
        })

    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'ok': False,
            'error': f'Fehler beim Generieren: {str(e)}'
        }), 500


@agent_bp.route('/analyze', methods=['POST'])
@login_required
def analyze_requirements():
    """Analyze requirements using AI"""
    try:
        data = request.get_json()
        requirements_text = data.get('requirements', '')
        
        if not requirements_text:
            return jsonify({'error': 'No requirements provided'}), 400
        
        # Initialize AI client
        ai_client = AIClient()
        
        # Analyze requirements
        analysis = ai_client.analyze_requirements(requirements_text)
        
        return jsonify({'analysis': analysis})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@agent_bp.route('/suggest/<int:requirement_id>', methods=['POST'])
@login_required
def suggest_improvements(requirement_id):
    """Suggest improvements for requirements"""
    try:
        requirement = Requirement.query.get_or_404(requirement_id)
        
        # Authorization check
        if requirement.project.user_id != current_user.id:
            abort(403)
        
        # Get latest version
        latest_version = requirement.get_latest_version()
        if not latest_version:
            return jsonify({'error': 'No version found for this requirement'}), 404
        
        # Initialize AI client
        ai_client = AIClient()
        
        # Get suggestions
        suggestions = ai_client.suggest_improvements(latest_version)
        
        return jsonify({'suggestions': suggestions})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
