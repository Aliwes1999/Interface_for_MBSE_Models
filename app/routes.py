from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify
from flask_login import login_required, current_user
from sqlalchemy import func, and_
import json
import os
from datetime import datetime
from . import db
from .models import Project, Requirement, RequirementVersion, ProjectFile
from .services.ai_client import generate_requirements

bp = Blueprint('main', __name__)

@bp.route("/")
@login_required
def home():
    projects = Project.query.filter_by(user_id=current_user.id).all()
    return render_template("start.html", projects=projects)

@bp.route("/create", methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        if project_name:
            # The concept of dynamic columns is removed in the new model.
            new_project = Project(name=project_name, user_id=current_user.id)
            db.session.add(new_project)
            db.session.commit()
        return redirect(url_for('main.home'))
    # The create.html is now a generic "new project" page if no project is passed.
    return render_template("create.html", project=None)

@bp.route("/project/<int:project_id>")
@login_required
def manage_project(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)
    
    # Show entry page with two options: generate or upload
    return render_template("project_entry.html", project=project)


@bp.route("/project/<int:project_id>/overview")
@login_required
def project_overview(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

    # Read explicit active_tab (e.g. 'files') from query string. If not provided
    # but a file_id is present, default to the requirements tab so the user can
    # immediately edit the imported/generated requirements.
    focus_file_id = request.args.get('file_id', type=int)
    active_tab = request.args.get('active_tab')
    if not active_tab and focus_file_id:
        active_tab = 'requirements'
    
    # Determine whether to show archive or requirements
    show_archive = (active_tab == 'files')

    # Get all requirements with ALL versions (not just the latest)
    # Filter out deleted requirements
    # If a specific file_id is provided, only show requirements from that file
    requirements = (
        Requirement.query
        .filter_by(project_id=project_id, is_deleted=False)
        .all()
    )
    
    # For each requirement, get versions (filtered by source_file_id if provided)
    req_with_versions = []
    generated_req_count = 0  # Count requirements from generated files (no upload source)
    
    for req in requirements:
        if focus_file_id:
            # Only include versions from the specific file
            versions = [v for v in req.versions if v.source_file_id == focus_file_id]
        else:
            # Show all versions
            versions = req.versions
        
        if versions:  # Only include requirements that have versions
            req_with_versions.append((req, versions))
            
            # Count requirements from generated files (source_file_id points to 'generated' type)
            if not focus_file_id:
                for v in versions:
                    if v.source_file_id:
                        source = ProjectFile.query.get(v.source_file_id)
                        if source and source.file_type == 'generated':
                            generated_req_count += 1
                            break  # Count each requirement only once
                    else:
                        # Old requirements without source_file_id are also considered generated
                        generated_req_count += 1
                        break
    
    # Extract ALL custom columns dynamically from requirement versions
    # This ensures we show all columns that exist in any requirement
    all_custom_columns = set()
    for req, versions in req_with_versions:
        for version in versions:
            custom_data = version.get_custom_data()
            if custom_data:
                all_custom_columns.update(custom_data.keys())
    
    # Convert to sorted list for consistent display
    custom_columns = sorted(list(all_custom_columns))
    
    # Get latest snapshot/export/generated file for quick access
    latest_snapshot = ProjectFile.query.filter_by(project_id=project_id).filter(
        ProjectFile.file_type.in_(['export', 'generated'])
    ).order_by(ProjectFile.created_at.desc()).first()
    
    # Get the source file name if viewing specific file
    source_file = None
    if focus_file_id:
        source_file = ProjectFile.query.get(focus_file_id)
    
    return render_template(
        "create.html", 
        project=project, 
        req_with_versions=req_with_versions,
        custom_columns=custom_columns,
        latest_snapshot=latest_snapshot,
        active_tab=active_tab,
        focus_file_id=focus_file_id,
        show_archive=show_archive,
        source_file=source_file,
        generated_req_count=generated_req_count
    )

@bp.route("/deleted_requirements")
@login_required
def deleted_requirements_overview():
    """Show all deleted requirements across all user's projects."""
    projects = Project.query.filter_by(user_id=current_user.id).all()
    
    # Collect deleted requirements from all projects
    all_deleted = []
    for project in projects:
        deleted_reqs = Requirement.query.filter_by(
            project_id=project.id, 
            is_deleted=True
        ).all()
        
        for req in deleted_reqs:
            latest_version = req.get_latest_version()
            if latest_version:
                all_deleted.append({
                    'project': project,
                    'requirement': req,
                    'version': latest_version
                })
    
    return render_template(
        "deleted_requirements_overview.html",
        deleted_items=all_deleted
    )

@bp.route("/requirement/<int:rid>/history")
@login_required
def requirement_history(rid):
    req = Requirement.query.get_or_404(rid)
    # Authorization check: ensure the user owns the project this requirement belongs to.
    if req.project.user_id != current_user.id:
        abort(403)
    
    # The 'versions' relationship is already ordered by version_index
    versions = req.versions
    return render_template("requirement_history.html", req=req, versions=versions)


@bp.route('/project/<int:project_id>/file/<int:file_id>')
@login_required
def view_project_file(project_id, file_id):
    """Show a detail page for a specific ProjectFile (preview columns / data)."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

    project_file = ProjectFile.query.get_or_404(file_id)
    if project_file.project_id != project.id:
        abort(404)

    preview = None
    columns = []
    try:
        # Only attempt to parse Excel uploads
        if project_file.filepath and project_file.filename.endswith(('.xlsx', '.xls')) and project_file.file_type == 'upload':
            preview = parse_excel_to_data(project_file.filepath)
            # preview expected as list of dicts; derive columns from keys of first row
            if preview and isinstance(preview, list) and len(preview) > 0:
                columns = list(preview[0].keys())
    except Exception:
        preview = None

    return render_template('file_view.html', project=project, file=project_file, preview=preview, columns=columns)
@bp.route('/file/<int:file_id>/download')
@login_required
def download_file(file_id):
    """Download a project file"""
    project_file = ProjectFile.query.get_or_404(file_id)
    project = project_file.project
    
    # Authorization check
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)
    
    # Check if file exists
    if not os.path.exists(project_file.filepath):
        abort(404)
    
    from flask import send_file
    return send_file(
        project_file.filepath,
        as_attachment=True,
        download_name=project_file.filename
    )

@bp.route('/file/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_project_file(file_id):
    """Delete a project file and its associated requirements"""
    project_file = ProjectFile.query.get_or_404(file_id)
    project = project_file.project
    
    # Authorization check - only project owner can delete files
    if project.user_id != current_user.id:
        abort(403)
    
    # Delete physical file
    if os.path.exists(project_file.filepath):
        try:
            os.remove(project_file.filepath)
        except Exception:
            pass  # Silently ignore file deletion errors
    
    # Delete all requirement versions associated with this file
    RequirementVersion.query.filter_by(source_file_id=file_id).delete()
    
    # Delete the ProjectFile entry
    db.session.delete(project_file)
    db.session.commit()
    
    flash(f"Datei '{project_file.filename}' wurde gelöscht.", "success")
    return redirect(url_for('main.project_overview', project_id=project.id, active_tab='files'))


@bp.route("/project/delete/<int:project_id>", methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    db.session.delete(project)
    db.session.commit()
    flash(f"Project '{project.name}' has been deleted.", "success")
    return redirect(url_for('main.home'))


# The routes below are now obsolete due to the data model refactoring
# and have been removed:
# - /project/<int:project_id>/deleted
# - /deleted_requirements_overview
# - /move/<int:project_id>/<int:req_id>/<string:from_table>/<string:to_table>
# - /edit/<int:project_id>/<int:req_id>
# - /export/<int:project_id>/<string:format>
# - /delete_column/<int:project_id>
# - /delete_requirement_permanently/<int:project_id>/<int:req_id>
# The simple /requirements and /add_requirement routes were also based on the old model.

# Routes for dynamic columns
@bp.route("/project/<int:project_id>/add_column", methods=['POST'])
@login_required
def add_column(project_id):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    column_name = request.form.get('column_name', '').strip()
    if not column_name:
        flash("Column name cannot be empty.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Get current columns and add the new one
    columns = project.get_custom_columns()
    if column_name in columns:
        flash(f"Column '{column_name}' already exists.", "warning")
    else:
        columns.append(column_name)
        project.set_custom_columns(columns)
        db.session.commit()
        flash(f"Column '{column_name}' added successfully.", "success")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

@bp.route("/project/<int:project_id>/remove_column/<column_name>", methods=['POST'])
@login_required
def remove_column(project_id, column_name):
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    # Get current columns and remove the specified one
    columns = project.get_custom_columns()
    if column_name in columns:
        columns.remove(column_name)
        project.set_custom_columns(columns)
        db.session.commit()
        flash(f"Column '{column_name}' removed successfully.", "success")
    else:
        flash(f"Column '{column_name}' not found.", "warning")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to update custom column data for a requirement version
@bp.route("/requirement_version/<int:version_id>/update_custom_data", methods=['POST'])
@login_required
def update_custom_data(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    if version.requirement.project.user_id != current_user.id:
        abort(403)
    
    column_name = request.form.get('column_name')
    value = request.form.get('value', '').strip()
    
    # Get current custom data and update it
    custom_data = version.get_custom_data()
    custom_data[column_name] = value
    version.set_custom_data(custom_data)
    db.session.commit()
    
    return jsonify({'success': True})

# Route to update requirement status
@bp.route("/requirement_version/<int:version_id>/update_status", methods=['POST'])
@login_required
def update_status(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    if version.requirement.project.user_id != current_user.id:
        abort(403)
    
    status = request.form.get('status')
    if status in ['Offen', 'In Arbeit', 'Fertig']:
        version.status = status
        db.session.commit()
        flash(f"Status updated to '{status}'.", "success")
    else:
        flash("Invalid status value.", "danger")
    
    return redirect(url_for('main.manage_project', project_id=version.requirement.project_id))

# AJAX route to get all versions of a requirement
@bp.route("/requirement/<int:req_id>/versions_json")
@login_required
def requirement_versions_json(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)
    
    versions_data = []
    for ver in req.versions:
        versions_data.append({
            'id': ver.id,
            'version_index': ver.version_index,
            'version_label': ver.version_label,
            'title': ver.title,
            'description': ver.description,
            'category': ver.category,
            'status': ver.status,
            'status_color': ver.get_status_color(),
            'custom_data': ver.get_custom_data(),
            'created_at': ver.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return jsonify(versions_data)

# Route to update requirement version data
@bp.route("/requirement_version/<int:version_id>/update", methods=['POST'])
@login_required
def update_requirement_version(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    if version.requirement.project.user_id != current_user.id:
        abort(403)
    
    # Get form data
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    category = request.form.get('category', '').strip()
    save_type = request.form.get('save_type', 'intermediate')  # 'intermediate' or 'final'
    
    # Validate required fields
    if not title or not description:
        flash("Title and description are required.", "danger")
        return redirect(url_for('main.manage_project', project_id=version.requirement.project_id))
    
    # Update fields
    version.title = title
    version.description = description
    version.category = category
    
    # Update status based on save type
    if save_type == 'intermediate':
        version.status = 'In Arbeit'
    elif save_type == 'final':
        version.status = 'Fertig'
    
    # Track who modified this version
    version.last_modified_by_id = current_user.id
    
    # Update custom data
    custom_data = version.get_custom_data()
    project = version.requirement.project
    custom_columns = project.get_custom_columns()
    
    for column in custom_columns:
        value = request.form.get(f'custom_{column}', '').strip()
        custom_data[column] = value
    
    version.set_custom_data(custom_data)
    
    # Save changes
    db.session.commit()
    
    flash(f"Anforderung erfolgreich aktualisiert. Status: {version.status}", "success")
    return redirect(request.referrer or url_for('main.manage_project', project_id=version.requirement.project_id))

# Route to delete a specific version of a requirement
@bp.route("/requirement_version/<int:version_id>/delete", methods=['POST'])
@login_required
def delete_requirement_version(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    req = version.requirement

    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)

    project_id = req.project_id

    # Check if there are any remaining versions
    remaining_versions = RequirementVersion.query.filter_by(requirement_id=req.id).count()

    if remaining_versions == 1:
        # This is the last version, mark the requirement as deleted instead of deleting the version
        req.is_deleted = True
        flash("Letzte Version gelöscht. Anforderung in Papierkorb verschoben.", "success")
    else:
        # Delete this specific version
        db.session.delete(version)
        flash(f"Version {version.version_label} erfolgreich gelöscht.", "success")

    db.session.commit()

    return redirect(request.referrer or url_for('main.manage_project', project_id=project_id))

# Route to soft delete a requirement (kept for compatibility, but marks all versions as deleted)
@bp.route("/requirement/<int:req_id>/delete", methods=['POST'])
@login_required
def delete_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)

    # Soft delete
    req.is_deleted = True
    db.session.commit()

    flash("Anforderung in Papierkorb verschoben.", "success")
    return redirect(request.referrer or url_for('main.manage_project', project_id=req.project_id))

# Route to restore a deleted requirement
@bp.route("/requirement/<int:req_id>/restore", methods=['POST'])
@login_required
def restore_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)
    
    # Restore
    req.is_deleted = False
    db.session.commit()
    
    flash("Requirement restored successfully.", "success")
    return redirect(url_for('main.deleted_requirements_overview'))

# Route to permanently delete a requirement
@bp.route("/requirement/<int:req_id>/delete_permanently", methods=['POST'])
@login_required
def delete_requirement_permanently(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)
    
    project_id = req.project_id
    
    # Permanently delete (cascade will delete all versions)
    db.session.delete(req)
    db.session.commit()
    
    flash("Requirement permanently deleted.", "success")
    return redirect(url_for('main.deleted_requirements_overview'))

# Route to regenerate a single requirement with AI
@bp.route("/requirement/<int:req_id>/regenerate", methods=['POST'])
@login_required
def regenerate_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    if req.project.user_id != current_user.id:
        abort(403)
    
    # Get the latest version to use as context
    latest_version = req.get_latest_version()
    if not latest_version:
        flash("No existing version found to regenerate.", "danger")
        return redirect(request.referrer or url_for('main.manage_project', project_id=req.project_id))
    
    try:
        # Get project's custom columns
        custom_columns = req.project.get_custom_columns()
        
        # Prepare context for AI
        context = {
            "project_name": req.project.name,
            "requirement_title": latest_version.title,
            "requirement_description": latest_version.description,
            "requirement_category": latest_version.category or "",
            "custom_data": latest_version.get_custom_data()
        }
        
        # Build complete columns list: title, description, custom columns, category
        columns = ["title", "description"] + custom_columns + ["category"]
        
        # Generate a new version with AI
        result = generate_single_requirement_alternative(context, columns)
        
        if not result:
            flash("Failed to generate alternative. AI returned empty result.", "danger")
            return redirect(request.referrer or url_for('main.manage_project', project_id=req.project_id))
        
        # Calculate next version
        next_index = latest_version.version_index + 1
        next_label = chr(ord('A') + (next_index - 1))
        
        # Create new version
        new_version = RequirementVersion(
            requirement_id=req.id,
            version_index=next_index,
            version_label=next_label,
            title=result.get("title", latest_version.title),
            description=result.get("description", latest_version.description),
            category=result.get("category", latest_version.category),
            status="Offen",  # New version starts as "Open"
            created_by_id=current_user.id,  # Track who created this version
            source_file_id=latest_version.source_file_id  # Keep same source file
        )
        
        # Get custom data from AI result or copy from previous version
        custom_data = {}
        for col in custom_columns:
            # Try to get value from AI result first, fallback to previous version
            value = result.get(col, latest_version.get_custom_data().get(col, ""))
            if value:
                custom_data[col] = value
        
        if custom_data:
            new_version.set_custom_data(custom_data)
        
        db.session.add(new_version)
        db.session.commit()
        
        flash(f"Neue Version {next_label} erfolgreich generiert!", "success")
        
    except Exception as e:
        flash(f"Fehler beim Generieren: {str(e)}", "danger")
    
    return redirect(request.referrer or url_for('main.manage_project', project_id=req.project_id))

def generate_single_requirement_alternative(context, columns):
    """Generate an alternative version of a requirement using AI."""
    try:
        # Prepare prompt for AI - explicitly ask for a DIFFERENT alternative
        prompt = f"""
        Erstelle eine ALTERNATIVE Version der folgenden Anforderung.
        
        WICHTIG: Die neue Version muss INHALTLICH UNTERSCHIEDLICH sein, aber das gleiche Ziel verfolgen.
        
        Projekt: {context['project_name']}
        
        Bisherige Version:
        Titel: {context['requirement_title']}
        Beschreibung: {context['requirement_description']}
        Kategorie: {context['requirement_category']}
        
        Zusätzliche Daten:
        {context['custom_data']}
        
        Erstelle eine neue Version, die:
        1. Einen anderen Ansatz oder eine andere Perspektive verfolgt
        2. Andere technische Details oder Spezifikationen enthält
        3. Eine alternative Formulierung oder Struktur hat
        4. Das gleiche Ziel erreicht, aber auf andere Weise
        
        Die neue Version sollte sich deutlich von der bisherigen unterscheiden, aber das gleiche Problem lösen.
        Sei kreativ und biete eine echte Alternative!
        """
        
        # Call the AI service
        ai_result = generate_requirements(prompt, {}, columns)
        
        # We expect a list of requirements, but we only need the first one
        if ai_result and len(ai_result) > 0:
            return ai_result[0]
        
        return None
        
    except Exception:
        raise

# Route to export project requirements to Excel
@bp.route("/project/<int:project_id>/export_excel")
@login_required
def export_excel(project_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    from flask import send_file
    
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    # Get all non-deleted requirements with their latest versions
    requirements = Requirement.query.filter_by(
        project_id=project_id,
        is_deleted=False
    ).all()
    
    # Extract custom columns dynamically from all requirements
    all_custom_columns = set()
    for req in requirements:
        latest_version = req.get_latest_version()
        if latest_version:
            custom_data = latest_version.get_custom_data()
            if custom_data:
                all_custom_columns.update(custom_data.keys())
    
    # Convert to sorted list for consistent display
    custom_columns = sorted(list(all_custom_columns))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    
    # Define headers: ID, title, description, category, dann custom columns (ohne diese 3), dann Version, Status
    # Filter custom_columns to exclude title, description, category
    filtered_custom_columns = [col for col in custom_columns if col not in ['title', 'description', 'category']]
    headers = ["ID", "title", "description", "category"] + filtered_custom_columns + ["Version", "Status"]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top")
    
    # Write data rows
    row_num = 2
    display_id = 1
    
    for req in requirements:
        latest_version = req.get_latest_version()
        if not latest_version:
            continue
        
        # Previously we filtered to only export requirements with status "Fertig".
        # Remove this filter so that all non-deleted requirements are exported
        # regardless of their status. If you want a filter, add a query
        # parameter and apply it before iterating requirements.
        
        custom_data = latest_version.get_custom_data()
        
        # Prepare row data: ID, title, description, category (from custom_data), custom columns, Version, Status
        row_data = [
            display_id,
            custom_data.get('title', '–'),
            custom_data.get('description', '–'),
            custom_data.get('category', '–')
        ]
        
        # Add other custom column values (excluding title, description, category)
        for col in filtered_custom_columns:
            row_data.append(custom_data.get(col, "–"))
        
        # Add version and status at the end
        row_data.append(latest_version.version_label)
        row_data.append(latest_version.status)
        
        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        row_num += 1
        display_id += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # ID
    ws.column_dimensions['B'].width = 30  # title
    ws.column_dimensions['C'].width = 50  # description
    ws.column_dimensions['D'].width = 20  # category
    
    # Set widths for custom columns (start at column E)
    col_letter_start = ord('E')
    for i, col in enumerate(filtered_custom_columns):
        col_letter = chr(col_letter_start + i)
        ws.column_dimensions[col_letter].width = 20
    
    # Set width for version and status (after custom columns)
    col_letter = chr(col_letter_start + len(filtered_custom_columns))
    ws.column_dimensions[col_letter].width = 10  # Version
    col_letter = chr(col_letter_start + len(filtered_custom_columns) + 1)
    ws.column_dimensions[col_letter].width = 15  # Status
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"requirements_{project.name.replace(' ', '_')}_{timestamp}.xlsx"

    # Save file to uploads directory (use absolute path)
    from flask import current_app
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
    os.makedirs(uploads_dir, exist_ok=True)
    filepath = os.path.join(uploads_dir, filename)
    wb.save(filepath)

    # Create database entry for exported file
    project_file = ProjectFile(
        project_id=project_id,
        filename=filename,
        filepath=filepath,
        file_type='export',
        created_by_id=current_user.id
    )
    db.session.add(project_file)
    db.session.commit()

    # Return file for download
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Route to import requirements from Excel
@bp.route("/project/<int:project_id>/import_excel", methods=['POST'])
@login_required
def import_excel(project_id):
    from openpyxl import load_workbook
    from werkzeug.utils import secure_filename
    import os
    
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    # Check if file was uploaded
    if 'excel_file' not in request.files:
        flash("Keine Datei ausgewählt.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    file = request.files['excel_file']
    
    if file.filename == '':
        flash("Keine Datei ausgewählt.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("Bitte laden Sie eine Excel-Datei (.xlsx oder .xls) hoch.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    try:
        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get custom columns for this project
        custom_columns = project.get_custom_columns()
        
        # Read header row to map columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Find column indices
        title_idx = None
        description_idx = None
        category_idx = None
        status_idx = None
        custom_col_indices = {}
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if header_lower in ['title', 'titel']:
                title_idx = idx
            elif header_lower in ['description', 'beschreibung']:
                description_idx = idx
            elif header_lower in ['category', 'kategorie']:
                category_idx = idx
            elif header_lower in ['status']:
                status_idx = idx
            elif header in custom_columns:
                custom_col_indices[header] = idx
        
        if title_idx is None or description_idx is None:
            flash("Excel-Datei muss mindestens 'Title' und 'Beschreibung' Spalten enthalten.", "danger")
            return redirect(url_for('main.manage_project', project_id=project_id))
        
        # Import rows (skip header)
        imported_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= title_idx:
                continue
            
            title = row[title_idx]
            if not title or str(title).strip() == '':
                continue
            
            title = str(title).strip()
            description = str(row[description_idx]).strip() if description_idx < len(row) and row[description_idx] else ""
            
            if not description:
                continue
            
            category = str(row[category_idx]).strip() if category_idx is not None and category_idx < len(row) and row[category_idx] else ""
            status = str(row[status_idx]).strip() if status_idx is not None and status_idx < len(row) and row[status_idx] else "Offen"
            
            # Validate status
            if status not in ['Offen', 'In Arbeit', 'Fertig']:
                status = 'Offen'
            
            # Create requirement
            from .agent import normalize_key
            key = normalize_key(title)
            
            req = Requirement.query.filter_by(project_id=project_id, key=key).first()
            
            if not req:
                req = Requirement(project_id=project_id, key=key)
                db.session.add(req)
                db.session.flush()
                version_index = 1
                version_label = 'A'
            else:
                # Create new version
                last_version = req.versions[-1] if req.versions else None
                version_index = last_version.version_index + 1 if last_version else 1
                version_label = chr(ord('A') + (version_index - 1))
            
            # Create version
            new_version = RequirementVersion(
                requirement_id=req.id,
                version_index=version_index,
                version_label=version_label,
                title=title,
                description=description,
                category=category,
                status=status,
                created_by_id=current_user.id
            )
            
            # Add custom column data
            custom_data = {}
            for col_name, col_idx in custom_col_indices.items():
                if col_idx < len(row) and row[col_idx]:
                    custom_data[col_name] = str(row[col_idx]).strip()
            
            if custom_data:
                new_version.set_custom_data(custom_data)
            
            db.session.add(new_version)
            imported_count += 1
        
        db.session.commit()
        flash(f"{imported_count} Anforderungen erfolgreich importiert!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Importieren: {str(e)}", "danger")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to share project with another user
@bp.route("/project/<int:project_id>/share", methods=['POST'])
@login_required
def share_project(project_id):
    from .models import User
    
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    email = request.form.get('email', '').strip()
    if not email:
        flash("Bitte geben Sie eine E-Mail-Adresse ein.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Find user by email
    user = User.query.filter_by(email=email).first()
    if not user:
        flash(f"Benutzer mit E-Mail '{email}' nicht gefunden.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    if user.id == current_user.id:
        flash("Sie können das Projekt nicht mit sich selbst teilen.", "warning")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Check if already shared
    if user in project.shared_with:
        flash(f"Projekt ist bereits mit {email} geteilt.", "warning")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Share project
    project.shared_with.append(user)
    db.session.commit()
    
    flash(f"Projekt erfolgreich mit {email} geteilt!", "success")
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to unshare project
@bp.route("/project/<int:project_id>/unshare/<int:user_id>", methods=['POST'])
@login_required
def unshare_project(project_id, user_id):
    from .models import User
    
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        abort(403)
    
    user = User.query.get_or_404(user_id)
    
    if user in project.shared_with:
        project.shared_with.remove(user)
        db.session.commit()
        flash(f"Projekt-Freigabe für {user.email} entfernt.", "success")
    else:
        flash("Benutzer hat keinen Zugriff auf dieses Projekt.", "warning")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to toggle requirement version block status
@bp.route("/requirement_version/<int:version_id>/toggle_block", methods=['POST'])
@login_required
def toggle_block_requirement(version_id):
    from datetime import datetime
    
    version = RequirementVersion.query.get_or_404(version_id)
    project = version.requirement.project
    
    # Authorization check - only project owner can block
    if project.user_id != current_user.id:
        abort(403)
    
    # Toggle block status
    if version.is_blocked:
        # Unblock
        version.is_blocked = False
        version.blocked_by_id = None
        version.blocked_at = None
        flash(f"Version {version.version_label} wurde freigegeben.", "success")
    else:
        # Block
        version.is_blocked = True
        version.blocked_by_id = current_user.id
        version.blocked_at = datetime.utcnow()
        flash(f"Version {version.version_label} wurde blockiert.", "warning")
    
    db.session.commit()
    return redirect(request.referrer or url_for('main.manage_project', project_id=project.id))

@bp.route("/requirement/version/<int:version_id>/update_quantifizierbar", methods=['POST'])
@login_required
def update_quantifizierbar(version_id):
    """Update quantifizierbar value for a requirement version"""
    version = RequirementVersion.query.get_or_404(version_id)
    project = version.requirement.project
    
    # Authorization check
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)
    
    # Get value from request
    data = request.get_json()
    quantifizierbar_value = data.get('quantifizierbar')
    
    if quantifizierbar_value not in ['ja', 'nein']:
        return jsonify({'ok': False, 'error': 'Invalid value'}), 400
    
    # Update custom_data
    custom_data = version.get_custom_data() or {}
    custom_data['quantifizierbar'] = quantifizierbar_value
    version.set_custom_data(custom_data)
    
    db.session.commit()
    
    return jsonify({'ok': True, 'value': quantifizierbar_value})

@bp.route("/hello")
def hello():
    return "Hello from Blueprint!"

