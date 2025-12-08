from flask import flash, redirect, url_for, send_file
from flask_login import current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from werkzeug.utils import secure_filename
import os
from .. import db
from ..models import Project, Requirement, RequirementVersion
from ..agent import normalize_key

def export_excel(project_id):
    """Export project requirements to Excel file."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        from flask import abort
        abort(403)

    # Get all non-deleted requirements with their latest versions
    requirements = Requirement.query.filter_by(
        project_id=project_id,
        is_deleted=False
    ).all()

    # Get custom columns
    custom_columns = project.get_custom_columns()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    # Define headers
    headers = ["Version", "ID", "Title", "Beschreibung"] + custom_columns + ["Kategorie", "Status"]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top")

    # Write data rows
    row_num = 2
    display_id = 1

    for req in requirements:
        latest_version = req.get_latest_version()
        if not latest_version:
            continue

        custom_data = latest_version.get_custom_data()

        # Prepare row data
        row_data = [
            latest_version.version_label,
            display_id,
            latest_version.title,
            latest_version.description
        ]

        # Add custom column values
        for col in custom_columns:
            row_data.append(custom_data.get(col, "–"))

        # Add category and status
        row_data.append(latest_version.category or "–")
        row_data.append(latest_version.status)

        # Write row
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_num += 1
        display_id += 1

    # Set column widths
    ws.column_dimensions['A'].width = 10  # Version
    ws.column_dimensions['B'].width = 8   # ID
    ws.column_dimensions['C'].width = 30  # Title
    ws.column_dimensions['D'].width = 50  # Description

    # Set widths for custom columns
    col_letter_start = ord('E')
    for i, col in enumerate(custom_columns):
        col_letter = chr(col_letter_start + i)
        ws.column_dimensions[col_letter].width = 20

    # Set widths for category and status
    col_letter = chr(col_letter_start + len(custom_columns))
    ws.column_dimensions[col_letter].width = 20  # Category
    col_letter = chr(col_letter_start + len(custom_columns) + 1)
    ws.column_dimensions[col_letter].width = 15  # Status

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create filename
    filename = f"requirements_{project.name.replace(' ', '_')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def import_excel(project_id, file):
    """Import requirements from Excel file."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id:
        from flask import abort
        abort(403)

    if file.filename == '':
        flash("Keine Datei ausgewählt.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))

    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("Bitte laden Sie eine Excel-Datei (.xlsx oder .xls) hoch.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))

    try:
        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Get custom columns for this project
        custom_columns = project.get_custom_columns()

        # Read header row to map columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # Find column indices
        title_idx = None
        description_idx = None
        category_idx = None
        status_idx = None
        custom_col_indices = {}

        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if header_lower in ['title', 'titel']:
                title_idx = idx
            elif header_lower in ['description', 'beschreibung']:
                description_idx = idx
            elif header_lower in ['category', 'kategorie']:
                category_idx = idx
            elif header_lower in ['status']:
                status_idx = idx
            elif header in custom_columns:
                custom_col_indices[header] = idx

        if title_idx is None or description_idx is None:
            flash("Excel-Datei muss mindestens 'Title' und 'Beschreibung' Spalten enthalten.", "danger")
            return redirect(url_for('main.manage_project', project_id=project_id))

        # Import rows (skip header)
        imported_count = 0
        new_versions = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= title_idx:
                continue

            title = row[title_idx]
            if not title or str(title).strip() == '':
                continue

            title = str(title).strip()
            description = str(row[description_idx]).strip() if description_idx < len(row) and row[description_idx] else ""

            if not description:
                continue

            category = str(row[category_idx]).strip() if category_idx is not None and category_idx < len(row) and row[category_idx] else ""
            status = str(row[status_idx]).strip() if status_idx is not None and status_idx < len(row) and row[status_idx] else "Offen"

            # Validate status
            if status not in ['Offen', 'In Arbeit', 'Fertig']:
                status = 'Offen'

            # Create requirement
            key = normalize_key(title)

            req = Requirement.query.filter_by(project_id=project_id, key=key).first()

            if not req:
                req = Requirement(project_id=project_id, key=key)
                db.session.add(req)
                db.session.flush()
                version_index = 1
                version_label = 'A'
            else:
                # Create new version
                last_version = req.versions[-1] if req.versions else None
                version_index = last_version.version_index + 1 if last_version else 1
                version_label = chr(ord('A') + (version_index - 1))

            # Create version
            new_version = RequirementVersion(
                requirement_id=req.id,
                version_index=version_index,
                version_label=version_label,
                title=title,
                description=description,
                category=category,
                status=status,
                created_by_id=current_user.id
            )

            # Add custom column data
            custom_data = {}
            for col_name, col_idx in custom_col_indices.items():
                if col_idx < len(row) and row[col_idx]:
                    custom_data[col_name] = str(row[col_idx]).strip()

            if custom_data:
                new_version.set_custom_data(custom_data)

            new_versions.append(new_version)
            imported_count += 1

        # Use bulk_save_objects for better performance
        if new_versions:
            db.session.bulk_save_objects(new_versions)

        db.session.commit()
        flash(f"{imported_count} Anforderungen erfolgreich importiert!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Importieren: {str(e)}", "danger")

    return redirect(url_for('main.manage_project', project_id=project_id))
