"""
Excel Service Module
Provides functions for parsing and processing Excel files.
"""

from openpyxl import load_workbook
from typing import List, Dict, Any, Optional
import os


def parse_excel_to_data(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parse Excel file and return data as list of dictionaries.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet to read. If None, reads the active sheet.
    
    Returns:
        List[Dict[str, Any]]: List of dictionaries containing the Excel data,
                              where keys are column headers and values are cell values.
    
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file format is invalid or cannot be read
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    if not file_path.endswith(('.xlsx', '.xls')):
        raise ValueError(f"Invalid file format. Expected .xlsx or .xls, got: {file_path}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        
        # Get worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Read header row - only include non-empty headers
        headers = []
        header_indices = []  # Track which column indices have valid headers
        for idx, cell in enumerate(ws[1]):
            if cell.value and str(cell.value).strip():
                headers.append(str(cell.value).strip())
                header_indices.append(idx)
        
        if not headers:
            raise ValueError("Excel file has no headers in the first row")
        
        # Read data rows
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # Skip completely empty rows
            
            row_dict = {}
            # Only process columns that have headers
            for header, col_idx in zip(headers, header_indices):
                if col_idx < len(row):
                    value = row[col_idx]
                    # Convert value to string and strip whitespace
                    if value is not None and str(value).strip():
                        row_dict[header] = str(value).strip()
                    else:
                        row_dict[header] = ""
                else:
                    row_dict[header] = ""
            
            # Only add row if it has at least one non-empty value (excluding system columns)
            if any(v for k, v in row_dict.items() if v and k.lower() not in ['version', 'id']):
                data.append(row_dict)
        
        return data
    
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")


def validate_excel_structure(file_path: str, required_columns: List[str]) -> tuple[bool, str]:
    """
    Validate that an Excel file has the required column structure.
    
    Args:
        file_path (str): Path to the Excel file
        required_columns (List[str]): List of required column names
    
    Returns:
        tuple[bool, str]: (is_valid, error_message)
                         Returns (True, "") if valid, (False, error_message) if invalid
    """
    try:
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Read headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        # Check for required columns (case-insensitive)
        missing_columns = []
        for required_col in required_columns:
            if required_col.lower() not in headers:
                missing_columns.append(required_col)
        
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"
        
        return True, ""
    
    except Exception as e:
        return False, f"Error validating Excel file: {str(e)}"


def get_excel_headers(file_path: str, sheet_name: Optional[str] = None) -> List[str]:
    """
    Get the header row from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str, optional): Name of the sheet to read. If None, reads the active sheet.
    
    Returns:
        List[str]: List of column headers
    
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file cannot be read
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        return headers
    
    except Exception as e:
        raise ValueError(f"Error reading Excel headers: {str(e)}")


def get_sheet_names(file_path: str) -> List[str]:
    """
    Get all sheet names from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        List[str]: List of sheet names
    
    Raises:
        FileNotFoundError: If the file doesn't exist
        ValueError: If the file cannot be read
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        return wb.sheetnames
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")
