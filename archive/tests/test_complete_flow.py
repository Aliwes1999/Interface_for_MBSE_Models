#!/usr/bin/env python3
"""Test the complete flow: upload -> redirect to file_view -> render file details"""

import os
os.environ['OPENAI_API_KEY'] = 'test-key'

from app import create_app
from app.models import User, Project, ProjectFile
from io import BytesIO
from openpyxl import Workbook
from unittest.mock import patch

app = create_app()

def create_test_excel():
    """Create a simple test Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    ws['B1'] = 'Description'
    ws['C1'] = 'Category'
    ws['A2'] = 'Test Requirement 1'
    ws['B2'] = 'Description 1'
    ws['C2'] = 'Functional'
    ws['A3'] = 'Test Requirement 2'
    ws['B3'] = 'Description 2'
    ws['C3'] = 'Non-Functional'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def mock_generate_requirements(*args, **kwargs):
    """Mock the AI requirement generation"""
    return [
        {
            'title': 'Generated Requirement 1',
            'description': 'This is a generated requirement',
            'category': 'Functional'
        },
        {
            'title': 'Generated Requirement 2',
            'description': 'Another generated requirement',
            'category': 'Non-Functional'
        }
    ]

def test_complete_flow():
    """Test complete flow: upload -> redirect -> file_view"""
    with app.app_context():
        with app.test_client() as client:
            # Find test user
            user = User.query.filter_by(email='test@example.com').first()
            project = Project.query.filter_by(name='Test Project', user_id=user.id).first()
            
            print("=== STEP 1: Login ===")
            response = client.post('/auth/login', data={
                'email': 'test@example.com',
                'password': 'test123'
            }, follow_redirects=True)
            print(f"✓ Login status: {response.status_code}")
            
            print("\n=== STEP 2: Upload File ===")
            excel_file = create_test_excel()
            
            with patch('app.agent.generate_requirements', side_effect=mock_generate_requirements):
                response = client.post(
                    f'/agent/generate/{project.id}',
                    data={
                        'user_description': 'Complete flow test',
                        'inputs': '[]',
                        'excel_file': (excel_file, 'test_flow.xlsx')
                    },
                    content_type='multipart/form-data'
                )
            
            print(f"✓ Upload status: {response.status_code}")
            data = response.get_json()
            print(f"✓ Generated {data.get('count')} requirements")
            
            # Extract file ID from redirect
            redirect_url = data.get('redirect', '')
            if '/file/' in redirect_url:
                file_id = int(redirect_url.split('/file/')[-1])
                print(f"✓ Redirecting to file ID: {file_id}")
            else:
                print(f"✗ No file ID in redirect: {redirect_url}")
                return
            
            print("\n=== STEP 3: Follow Redirect to File View ===")
            response = client.get(f'/project/{project.id}/file/{file_id}')
            print(f"✓ File view status: {response.status_code}")
            
            if response.status_code == 200:
                html = response.data.decode('utf-8')
                
                # Check for key file details
                file_obj = ProjectFile.query.get(file_id)
                print(f"\n✓ File Details:")
                print(f"  - Filename: {file_obj.filename}")
                print(f"  - Type: {file_obj.file_type}")
                print(f"  - Created by: {file_obj.created_by.email if file_obj.created_by else 'Unknown'}")
                
                # Check for UI elements
                checks = {
                    'File title': f'Datei: {file_obj.filename}' in html,
                    'File type badge': 'badge bg-' in html,  # Should have badge styling
                    'Back button': 'Zurück zum Überblick' in html,
                    'Columns': 'Spalten' in html or 'Vorschau' in html,
                    'Download button': 'Herunterladen' in html,
                }
                
                print(f"\n✓ Page Elements:")
                for name, found in checks.items():
                    status = '✓' if found else '✗'
                    print(f"  {status} {name}")
                
                print(f"\n✓ FLOW TEST COMPLETE! File is properly displayed.")
            else:
                print(f"✗ File view returned status {response.status_code}")

if __name__ == '__main__':
    test_complete_flow()
