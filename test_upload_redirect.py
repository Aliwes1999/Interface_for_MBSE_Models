#!/usr/bin/env python3
"""Test script to simulate a file upload and verify the redirect to file_view"""

from app import create_app
from app.models import User, Project, ProjectFile
from werkzeug.datastructures import FileStorage
from io import BytesIO
import json
from openpyxl import Workbook

app = create_app()

def create_test_excel():
    """Create a simple test Excel file"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    ws['B1'] = 'Description'
    ws['C1'] = 'Category'
    ws['A2'] = 'Test Requirement'
    ws['B2'] = 'This is a test'
    ws['C2'] = 'Functional'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def test_upload_redirect():
    """Test that file upload redirects to file_view"""
    with app.app_context():
        with app.test_client() as client:
            # Find test user
            user = User.query.filter_by(email='test@example.com').first()
            if not user:
                print("✗ Test user not found")
                return
            
            print(f"✓ Test user found: {user.email}")
            
            # Find test project
            project = Project.query.filter_by(name='Test Project', user_id=user.id).first()
            if not project:
                print("✗ Test project not found")
                return
            
            print(f"✓ Test project found: {project.name} (ID: {project.id})")
            
            # Login first
            response = client.post('/auth/login', data={
                'email': 'test@example.com',
                'password': 'test123'
            }, follow_redirects=True)
            
            print(f"\nLogin response status: {response.status_code}")
            
            # Now try to upload a file
            print(f"\nTesting file upload to /agent/generate/{project.id}...")
            
            excel_file = create_test_excel()
            
            response = client.post(
                f'/agent/generate/{project.id}',
                data={
                    'user_description': 'Test upload',
                    'inputs': '[]',
                    'excel_file': (excel_file, 'test_upload.xlsx')
                },
                content_type='multipart/form-data'
            )
            
            print(f"Upload response status: {response.status_code}")
            
            if response.status_code == 200:
                try:
                    data = response.get_json()
                    print(f"✓ Response JSON: {json.dumps(data, indent=2)}")
                    
                    if 'redirect' in data:
                        print(f"\n✓ Redirect URL: {data['redirect']}")
                        
                        # Check if it redirects to file_view
                        if 'file' in data['redirect']:
                            print(f"✓ Correctly redirects to file view!")
                        else:
                            print(f"✗ Does not redirect to file view")
                except:
                    print(f"Response body: {response.data}")
            else:
                print(f"Response body: {response.data}")

if __name__ == '__main__':
    test_upload_redirect()
